import pandas as pd
import os
from openpyxl import load_workbook

dir  = r'D:\Meena\Automation\correlation\Fw Updated 7P5T ABB RO Sim data with Code and VREF'
outpath = r"D:\Meena\Automation\correlation"
out_fname = "7p5tSIMfull.xlsx"
if not os.path.exists(outpath):
    os.makedirs(outpath)
outpath1 = os.path.join(outpath, out_fname)
writer = pd.ExcelWriter(outpath1, 'xlsxwriter')

df_out=pd.DataFrame()
for path, subdirs, files in os.walk(dir):
        for name in files:
                fname = (os.path.join(path, name))
                print(fname)
                wb = load_workbook(fname, read_only=True)
                print(wb.sheetnames)
                for j in wb.sheetnames:
                    df = pd.read_excel(fname, sheet_name=j)
                    df_out = pd.concat([df_out, df], axis=0,sort=False)
                    print(df_out.shape)
        df_out['Cell_name'] = df_out['RO Name'].str.split('_').str[-1]
        df_out['Track'] = df_out['RO Name'].str.split('_INV').str[0].str.split('_').str[-1]
        df_out.to_excel(writer, index=False)


writer.save()
writer.close()