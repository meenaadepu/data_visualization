import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
import os
from openpyxl import load_workbook
import numpy as np
sns.set(style='whitegrid',)

folder = "ch0"
inputpath = r"D:\Meena\Automation\ATV\2227\tsvs\plots"
outpath = r"D:\Meena\Automation\ATV\2227\tsvs\outputs"
out_path = os.path.join(outpath, folder)
Spec_Max=0.01
Spec_Min=-0.01
if not os.path.exists(out_path):
    os.makedirs(out_path)
df_total = pd.DataFrame()
for path, subdirs, files in os.walk(inputpath):
    for name in files:
        print(name)
        fname = (os.path.join(path, name))
        wb = load_workbook(fname, read_only=True)
        sheets = list(wb.sheetnames)
        for i in sheets:
            df = pd.read_excel(fname,sheet_name=i, header=[2,3,4])
           # print(df.columns)
            df.columns = df.columns.map('_'.join)
            df.rename_axis('applied voltage').reset_index()
            df1 = pd.concat([(df.iloc[:, 1:2]), (df.iloc[:, 18:])], axis=1)
            df_total = pd.concat([df_total,df1],axis=1)
            print(df_total.shape)
            cols= list(df1.columns)[1:]
            f1 = plt.figure(figsize=(18, 7))
            print(cols)
            for i in cols:
               ax1 = f1.add_subplot(111)
               sns.lineplot(data=df1, x=df1.iloc[:,0], y=i,ax=ax1,marker='o')
            Title = str(i).split('-')[0].split('_')[1]+ str(name).split('_')[2].split('.')[0]
            ax1.set_xlim([0.4, 1.4])
            ax1.set_ylim([-0.015,0.015])
            ax1.axhline(y=Spec_Max, linestyle='-.', color='r', label="Spec_Max")
            ax1.axhline(y=Spec_Min, linestyle='--', color='b', label="Spec_Min")
            ax1.set_xlabel("Applied Input Voltage CH0(V)", size=12.5, fontdict=dict(weight='bold'), labelpad=10)
            ax1.set_ylabel("Output Voltage Accuracy(V)", size=12.5, fontdict=dict(weight='bold'), labelpad=10)
            ax1.set_title(Title, size=12.5, fontdict=dict(weight='bold'))
            ax1.legend(loc='upper center', bbox_to_anchor=(0.5, -0.19), fancybox=False, shadow=False, ncol=10,
                       edgecolor='black', prop={'size': 12})
            for axis in ['top', 'bottom', 'left', 'right']:
                ax1.spines[axis].set_linewidth(2)
                ax1.spines[axis].set_color("black")
            print(Title)
            f1_outpath = out_path + "\\" + Title + ".png"
            f1.savefig(f1_outpath, bbox_inches='tight')



df_total = df_total.loc[:, ~df_total.columns.duplicated()]
df_total.to_csv("ch0df_total.csv")
cols_total = list(df_total.columns)[1:]
print(cols_total)
f2 = plt.figure(figsize=(18, 7))
ax2 = f2.add_subplot(111)
for j in cols_total:
    Title1 = "VS Accuracy Measurement"
    sns.lineplot(data=df_total, x=df_total.iloc[:, 0], y=j,ax=ax2,marker='o')
ax2.set_xlim([0.4, 1.4])
ax2.set_ylim([-0.015,0.015])
ax2.axhline(y=Spec_Max, linestyle='-.', color='r', label="Spec_Max")
ax2.axhline(y=Spec_Min, linestyle='--', color='b', label="Spec_Min")
ax2.set_xlabel("Applied Input Voltage CH0(V)", size=12.5, fontdict=dict(weight='bold'), labelpad=10)
ax2.set_ylabel("Output Voltage Accuracy(V)", size=12.5, fontdict=dict(weight='bold'), labelpad=10)
ax2.set_title( Title1, size=12.5, fontdict=dict(weight='bold'))
ax2.legend(loc='upper center', bbox_to_anchor=(0.5, -0.19), fancybox=False, shadow=False, ncol=10,
               edgecolor='black',prop={'size': 12})
for axis in ['top', 'bottom', 'left', 'right']:
    ax2.spines[axis].set_linewidth(2)
    ax2.spines[axis].set_color("black")
#plt.show()
df_total.to_csv("df_total.csv")
f2_outpath = out_path + "\\" +Title1  + ".png"
f2.savefig(f2_outpath, bbox_inches='tight')
plt.cla()
print("completed")





