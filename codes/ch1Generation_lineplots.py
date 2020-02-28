import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
import os
from openpyxl import load_workbook
from collections import OrderedDict
import numpy as np

folder = "ch1"
inputpath = r"D:\Meena\Automation\ATV\2227\tsvs\ch1"
outpath = r"D:\Meena\Automation\ATV\2227\tsvs\outputs"
out_path = os.path.join(outpath, folder)
Spec_Max=0.01
Spec_Min=-0.01
if not os.path.exists(out_path):
    os.makedirs(out_path)
df_total = pd.DataFrame()
for path, subdirs, files in os.walk(inputpath):
    for name in files:
        print(name)
        fname = (os.path.join(path, name))
        wb = load_workbook(fname, read_only=True)
        sheets = list(wb.sheetnames)
        for i in sheets:
            df = pd.read_excel(fname,sheet_name=i, header=[3,4,5])
           # print(df.columns)
            df.columns = df.columns.map('_'.join)
            df.rename_axis('i/p voltage').reset_index()
            df1 = pd.concat([(df.iloc[:, 1:2]), (df.iloc[:, 48:])], axis=1)
            df_total = pd.concat([df_total,df1],axis=1)
df_total = df_total.loc[:, ~df_total.columns.duplicated()]
df_total = df_total.rename(columns={col: col.split('.')[0] for col in df.columns})
cols_total = list(df_total.columns)[1:]
print(cols_total)
f2 = plt.figure(figsize=(12, 7))
ax2 = f2.add_subplot(111)
ax2.grid(True)
for j in cols_total:
    sns.set(style="ticks")
    Title1 = "Voltage Sensor Accuracy Measurement CH1"
    g=sns.lineplot(data=df_total, x=df_total.iloc[:, 0], y=j,ax=ax2,linewidth=3,marker='o')
    g.legend(cols_total)
ax2.text(0.5, Spec_Min, "Spec_Min", va='bottom', ha="center",color="k", transform=ax2.get_yaxis_transform(), size=14,)
ax2.text(0.5, Spec_Max, "Spec_Max", va='bottom', ha="center",color="k", transform=ax2.get_yaxis_transform(), size=14,)
ax2.axhline(y=Spec_Max, linestyle='--', color='k',linewidth=2.5,label ="Spec_Max")
ax2.axhline(y=Spec_Min, linestyle='--', color='k',linewidth=2.5,label ="Spec_Min")
g.legend(cols_total,loc='upper center', bbox_to_anchor=(0.5, -0.2),fancybox=False, shadow=False, ncol=10, edgecolor='black',fontsize=5.5)

ax2.set_xlabel("Applied Input Voltage CH1(V)", size=12.5, fontdict=dict(weight='bold'), labelpad=10)
ax2.set_ylabel("IP Output Voltage Accuracy(V)", size=12.5, fontdict=dict(weight='bold'), labelpad=10)
ax2.set_title( Title1, size=13.5, fontdict=dict(weight='bold'))
ax2.tick_params(which='major', length=5, width=4, labelsize=14.5)
ax2.set_xlim([0.4, 1.4])
ax2.set_ylim([-0.015,0.015])
# ax2.set_yticks(np.arange(-6, 7, 1))
# ax2.set_xticks(np.arange(-40, 130, 10))
for axis in ['top', 'bottom', 'left', 'right']:
    ax2.spines[axis].set_linewidth(2.5)
    ax2.spines[axis].set_color("black")

f2_outpath = out_path + "\\" +Title1  + ".png"
f2.savefig(f2_outpath, bbox_inches='tight')
plt.cla()
print("completed")








